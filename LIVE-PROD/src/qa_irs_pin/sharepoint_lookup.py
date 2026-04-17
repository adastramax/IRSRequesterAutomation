"""SharePoint workbook lookup for TEID -> state and site_name -> TEID.

Uses Microsoft Graph ROPC password grant. Credentials are optional — all
functions return None gracefully when credentials are absent or the lookup
fails, so processing is never blocked.
"""

from __future__ import annotations

import logging
import re
from datetime import date
from typing import Any

import requests

logger = logging.getLogger(__name__)

_CACHE: dict[str, Any] = {}

_SHAREPOINT_SITE = "adastrainccom.sharepoint.com:/sites/AdAstraInternalAdminstration-ReportCenter"
_WORKBOOK_NAME = "IRS Site Listing updated.xlsx"
_SHEET_NAME = "Master Site and State"

# TEIDs with known duplicate entries in the workbook — treat as ambiguous
_AMBIGUOUS_TEIDS: set[str] = set()


def _get_token(tenant: str, client_id: str, username: str, password: str) -> str | None:
    try:
        resp = requests.post(
            f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
            data={
                "grant_type": "password",
                "client_id": client_id,
                "scope": "https://graph.microsoft.com/.default",
                "username": username,
                "password": password,
            },
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()["access_token"]
    except Exception as exc:
        logger.warning("SharePoint Graph token request failed: %s", exc)
        return None


def _download_workbook(token: str) -> bytes | None:
    session = requests.Session()
    session.headers["Authorization"] = f"Bearer {token}"
    try:
        site_resp = session.get(f"https://graph.microsoft.com/v1.0/sites/{_SHAREPOINT_SITE}", timeout=30)
        site_resp.raise_for_status()
        site_id = site_resp.json()["id"]

        search_resp = session.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/search(q='{_WORKBOOK_NAME}')",
            timeout=30,
        )
        search_resp.raise_for_status()
        items = search_resp.json().get("value", [])
        if not items:
            logger.warning("SharePoint workbook '%s' not found.", _WORKBOOK_NAME)
            return None

        item = items[0]
        drive_id = item["parentReference"]["driveId"]
        item_id = item["id"]
        content_resp = session.get(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
            timeout=120,
        )
        content_resp.raise_for_status()
        return content_resp.content
    except Exception as exc:
        logger.warning("SharePoint workbook download failed: %s", exc)
        return None


def _normalize_site_key(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def _load_lookup_tables(
    tenant: str, client_id: str, username: str, password: str
) -> tuple[dict[str, dict], dict[str, dict]]:
    """Return (teid_map, site_name_map). Both keyed by normalized strings."""
    import io
    from openpyxl import load_workbook as _load_wb

    cache_key = f"sp_lookup_{date.today().isoformat()}"
    if cache_key in _CACHE:
        return _CACHE[cache_key]

    token = _get_token(tenant, client_id, username, password)
    if not token:
        return {}, {}

    wb_bytes = _download_workbook(token)
    if not wb_bytes:
        return {}, {}

    try:
        wb = _load_wb(filename=io.BytesIO(wb_bytes), read_only=True, data_only=True)
        if _SHEET_NAME not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook.", _SHEET_NAME)
            return {}, {}

        ws = wb[_SHEET_NAME]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return {}, {}

        header_map = {str(v).strip().lower(): i for i, v in enumerate(headers) if v is not None}
        site_col = header_map.get("site name")
        teid_col = header_map.get("te id")
        state_col = header_map.get("state")
        if any(c is None for c in (site_col, teid_col, state_col)):
            logger.warning("Expected columns missing from SharePoint sheet.")
            return {}, {}

        teid_map: dict[str, dict] = {}
        site_name_map: dict[str, dict] = {}
        site_name_teid_count: dict[str, int] = {}

        for row in rows_iter:
            raw_site = row[site_col] if len(row) > site_col else None
            raw_teid = row[teid_col] if len(row) > teid_col else None
            raw_state = row[state_col] if len(row) > state_col else None

            site_name = str(raw_site).strip() if raw_site else ""
            teid_str = str(int(raw_teid)).zfill(4) if isinstance(raw_teid, (int, float)) and raw_teid else str(raw_teid).strip() if raw_teid else ""
            state = str(raw_state).strip() if raw_state and str(raw_state).strip().lower() not in ("", "nan", "none") else ""

            if not teid_str or not site_name:
                continue

            entry = {"site_name": site_name, "teid": teid_str, "state": state}
            teid_map[teid_str] = entry

            site_key = _normalize_site_key(site_name)
            site_name_teid_count[site_key] = site_name_teid_count.get(site_key, 0) + 1
            site_name_map[site_key] = entry

        # Mark ambiguous site name keys (multiple different TEIDs)
        ambiguous_keys = {k for k, count in site_name_teid_count.items() if count > 1}
        for k in ambiguous_keys:
            site_name_map.pop(k, None)

        _CACHE[cache_key] = (teid_map, site_name_map)
        logger.info(
            "SharePoint lookup loaded: %d TEIDs, %d site names (%d ambiguous removed).",
            len(teid_map), len(site_name_map), len(ambiguous_keys),
        )
        return teid_map, site_name_map

    except Exception as exc:
        logger.warning("SharePoint workbook parse failed: %s", exc)
        return {}, {}


def _get_tables(config: Any | None = None) -> tuple[dict, dict]:
    """Load tables using env vars or provided config object."""
    import os
    tenant = (getattr(config, "GRAPH_TENANT", None) or os.getenv("GRAPH_TENANT", "")).strip()
    client_id = (getattr(config, "GRAPH_CLIENT_ID", None) or os.getenv("GRAPH_CLIENT_ID", "")).strip()
    username = (getattr(config, "GRAPH_USERNAME", None) or os.getenv("GRAPH_USERNAME", "")).strip()
    password = (getattr(config, "GRAPH_PASSWORD", None) or os.getenv("GRAPH_PASSWORD", "")).strip()

    if not all((tenant, client_id, username, password)):
        return {}, {}

    return _load_lookup_tables(tenant, client_id, username, password)


def get_state_for_teid(teid: str, config: Any | None = None) -> str | None:
    """Return 2-letter state abbreviation for a TEID, or None if not found."""
    if not teid:
        return None
    normalized = str(teid).strip().zfill(4)
    teid_map, _ = _get_tables(config)
    entry = teid_map.get(normalized)
    if entry and entry.get("state"):
        return entry["state"]
    return None


_US_STATE_ABBREVS = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY","DC",
}

def extract_state_from_site_name(site_name: str) -> str | None:
    """Try to extract a 2-letter US state abbreviation from a site name string.

    Handles patterns like:
      '500 Woodward Avenue, Detroit, MI, USA'  -> 'MI'
      'Silver Spring, MD, USA'                 -> 'MD'
      '801 Broad Street, Nashville, TN, USA'   -> 'TN'
    """
    if not site_name:
        return None
    # Look for a standalone 2-letter uppercase state code in the string
    parts = [p.strip().rstrip(".") for p in re.split(r"[,\s]+", site_name)]
    for part in reversed(parts):
        if part.upper() in _US_STATE_ABBREVS:
            return part.upper()
    return None


def get_teid_for_site_name(site_name: str, config: Any | None = None) -> dict | None:
    """Return {teid, site_name, state} for a site name, or None if not found / ambiguous."""
    if not site_name:
        return None
    _, site_name_map = _get_tables(config)
    if not site_name_map:
        return None

    key = _normalize_site_key(site_name)
    if key in site_name_map:
        return site_name_map[key]

    # Progressive fallback: strip parenthetical address suffix
    stripped_parens = re.sub(r"\s*\(.*?\)\s*$", "", site_name).strip()
    if stripped_parens and stripped_parens != site_name:
        key2 = _normalize_site_key(stripped_parens)
        if key2 in site_name_map:
            return site_name_map[key2]

    # Further strip trailing ", STATE"
    stripped_state = re.sub(r",\s*[A-Z]{2}\s*$", "", stripped_parens or site_name).strip()
    if stripped_state and stripped_state not in (site_name, stripped_parens):
        key3 = _normalize_site_key(stripped_state)
        if key3 in site_name_map:
            return site_name_map[key3]

    return None
